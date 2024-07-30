import openpyxl
import subprocess
import os
from openai import OpenAI
import sys

#your openai api key (second argument)
client = OpenAI(api_key=sys.argv[2])


#path of the excel file (first argument)
inputXlPath = sys.argv[1]

#loading the excel file
wb = openpyxl.load_workbook(inputXlPath)
sheet = wb.active

#array containing all the project paths from the excel file
path = []

#scanning the excel file for paths
i = 1
bool = True
while bool:
    tab = "A" + str(i)
    projectPath = sheet[tab].value
    i = i + 1
    if projectPath == None:
        bool = False
    else:
        path.append(projectPath)

#an array containing the outputs of the syft cli
cliOutputs = []

#running git clone and syft for each path
for p in path:
    tmp_clone_folder_path = os.path.expanduser('~') + "/tmp/" + p.split("/")[-1]
    os.makedirs(u)
    subprocess.run(["git", "clone", p, tmp_clone_folder_path])
    result = subprocess.check_output(['syft', tmp_clone_folder_path])
    cliOutputs.append(str(result.decode('utf-8')))
    subprocess.run(["rm", "-r", "-f", tmp_clone_folder_path])

#a new array containing the cleaned cli outputs of syft
cleanedList = []

#cleaning each cli output of syft and adding it to cleanedList array
for o in cliOutputs:
    raw_dependecies_arr = o.split("\n")
    raw_dependecies_arr.pop(0)
    raw_dependecies_arr.pop(-1)

    fixed_dependecies_arr = []

    for s in raw_dependecies_arr:
        is_no_name = False
        s = s.split(" ")
        if s[0] == "":
            is_no_name = True
        s = [i for i in s if i != '']
        if is_no_name == True:
            s.insert(0, "")
        fixed_dependecies_arr.append(s)

    cleanedList.append(fixed_dependecies_arr)

#a new array with the no name dependenceies and the "(+1 Duplicates)" text removed
removedList = []

#removing the no name dependenceies and the "(+1 Duplicates)" text
for l in cleanedList:
    cleaned_dependecies_project_arr = []
    for cleaned_dependecy_arr in l:
        if cleaned_dependecy_arr != []:
            if cleaned_dependecy_arr[0] != "":
                if cleaned_dependecy_arr[-1].endswith('duplicates)') or cleaned_dependecy_arr[-1].endswith('duplicate)'):
                    cleaned_dependecy_arr = cleaned_dependecy_arr[:-2]
                cleaned_dependecies_project_arr.append(cleaned_dependecy_arr)
    removedList.append(cleaned_dependecies_project_arr)

#an array containing chatgpt's descreptions
finalList = []

#adding to each dependency array chatgpt's descreption
for r in removedList:
    arr = []
    for dependency_arr in r:
        
        message = "Can you explain what the " + dependency_arr[-1] + " " + dependency_arr[0] + " dependency is doing?"

        stream = client.chat.completions.create(
           model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": message}],
           stream=True,
        )
        for part in stream:
               reply = part.choices[0].delta.content or ""

        dependency_arr.append(reply)
        arr.append(dependency_arr)
    finalList.append(arr)

#creating the output excel file
try:
    filepath_output = sys.argv[3]
except:
    filepath_output = os.path.expanduser('~') + "/output.xlsx"

wb_output = openpyxl.Workbook()

ws2 = wb_output.create_sheet(title = 'Output')
ws2 = wb_output.get_sheet_by_name('Output')

#adding the results back to excel
h = 0
for f in finalList:
    projectName = path[h].split("/")[-1]
    h = h + 1
    b = h
    for f1 in f:
        ws2['A' + str(h)] = projectName
        n = 2
        for k in f1:
            if n==2:
                ws2['B' + str(b)] = k
            if n==3:
                ws2['C' + str(b)] = k
            if n==4:
                ws2['D' + str(b)] = k
            if n==5:
                ws2['E' + str(b)] = k
            n = n + 1
        b = b + 1

#saving the excel file
wb_output.save(filepath_output)